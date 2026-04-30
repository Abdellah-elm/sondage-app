import uuid
from django.db import models
from django.utils import timezone
from django.utils.text import slugify
from django.contrib.auth.hashers import make_password, check_password
from django.http import HttpResponse
import csv
import io


class Sondage(models.Model):
    THEME_DEFAUT = 'defaut'
    THEME_MODERNE = 'moderne'
    THEME_PRO = 'professionnel'
    THEME_CHALEUREUX = 'chaleureux'
    THEMES = [
        (THEME_DEFAUT, 'Défaut (Bleu)'),
        (THEME_MODERNE, 'Moderne (Violet)'),
        (THEME_PRO, 'Professionnel (Vert)'),
        (THEME_CHALEUREUX, 'Chaleureux (Orange)'),
    ]

    createur = models.ForeignKey(
        'accounts.Utilisateur', on_delete=models.CASCADE, related_name='sondage_set'
    )
    titre = models.CharField(max_length=300)
    description = models.TextField(blank=True)
    slug = models.SlugField(max_length=350, unique=True)
    est_actif = models.BooleanField(default=True)
    est_anonyme = models.BooleanField(default=False)
    est_modele = models.BooleanField(default=False)
    est_archive = models.BooleanField(default=False)
    _etait_actif_avant_archive = models.BooleanField(default=True)
    date_limite = models.DateTimeField(null=True, blank=True)
    duree_limite_minutes = models.IntegerField(null=True, blank=True)
    max_reponses = models.IntegerField(null=True, blank=True)
    autoriser_integration = models.BooleanField(default=False)
    mot_de_passe = models.CharField(max_length=128, blank=True)
    theme = models.CharField(max_length=20, choices=THEMES, default=THEME_DEFAUT)
    cree_le = models.DateTimeField(auto_now_add=True)
    modifie_le = models.DateTimeField(auto_now=True)
    archive_le = models.DateTimeField(null=True, blank=True)

    class Meta:
        verbose_name = 'Sondage'
        verbose_name_plural = 'Sondages'
        ordering = ['-cree_le']

    def __str__(self):
        return self.titre

    def save(self, *args, **kwargs):
        if not self.slug:
            base_slug = slugify(self.titre)
            slug = base_slug
            n = 1
            while Sondage.objects.filter(slug=slug).exclude(pk=self.pk).exists():
                slug = f'{base_slug}-{n}'
                n += 1
            self.slug = slug
        super().save(*args, **kwargs)

    def est_expire(self):
        if self.date_limite:
            return timezone.now() > self.date_limite
        return False

    def a_minuteur(self):
        return self.duree_limite_minutes is not None

    def est_protege_par_mdp(self):
        return bool(self.mot_de_passe)

    def verifier_mot_de_passe(self, mdp_saisi):
        return check_password(mdp_saisi, self.mot_de_passe)

    def definir_mot_de_passe(self, mdp_brut):
        self.mot_de_passe = make_password(mdp_brut)

    def obtenir_nombre_reponses(self):
        return self.soumission_set.filter(est_complete=True).count()

    def obtenir_code_integration(self):
        if not self.autoriser_integration:
            return ''
        lien = self.liens.filter(est_actif=True, est_public=True).first()
        if lien:
            url = lien.obtenir_url_absolue()
            return f'<iframe src="{url}" width="100%" height="600" frameborder="0"></iframe>'
        return ''

    def dupliquer(self):
        nouveau = Sondage.objects.create(
            createur=self.createur,
            titre=f'Copie de {self.titre}',
            description=self.description,
            est_anonyme=self.est_anonyme,
            duree_limite_minutes=self.duree_limite_minutes,
            max_reponses=self.max_reponses,
        )
        for section in self.sections.order_by('ordre'):
            nouvelle_section = SectionSondage.objects.create(
                sondage=nouveau,
                titre=section.titre,
                description=section.description,
                ordre=section.ordre,
            )
            for question in section.questions.order_by('ordre'):
                nouvelle_question = Question.objects.create(
                    section=nouvelle_section,
                    texte=question.texte,
                    type_question=question.type_question,
                    est_obligatoire=question.est_obligatoire,
                    ordre=question.ordre,
                    echelle_min=question.echelle_min,
                    echelle_max=question.echelle_max,
                    libelle_min=question.libelle_min,
                    libelle_max=question.libelle_max,
                )
                for choix in question.choix.order_by('ordre'):
                    Choix.objects.create(
                        question=nouvelle_question,
                        texte=choix.texte,
                        ordre=choix.ordre,
                    )
        # LienPartage et StatistiqueSondage créés automatiquement par le signal
        return nouveau

    def archiver(self):
        self._etait_actif_avant_archive = self.est_actif
        self.est_archive = True
        self.est_actif = False
        self.archive_le = timezone.now()
        self.save()

    def desarchiver(self):
        self.est_archive = False
        self.est_actif = self._etait_actif_avant_archive
        self.archive_le = None
        self.save()

    def exporter_csv(self):
        response = HttpResponse(content_type='text/csv; charset=utf-8')
        response['Content-Disposition'] = f'attachment; filename="sondage_{self.slug}.csv"'
        response.write('﻿')  # BOM UTF-8 pour Excel
        writer = csv.writer(response)

        questions = Question.objects.filter(
            section__sondage=self
        ).order_by('section__ordre', 'ordre')

        entetes = ['Date soumission', 'Durée (min)'] + [q.texte for q in questions]
        writer.writerow(entetes)

        for soumission in self.soumission_set.filter(est_complete=True).order_by('termine_le'):
            ligne = [
                soumission.termine_le.strftime('%d/%m/%Y %H:%M') if soumission.termine_le else '',
                round(soumission.obtenir_duree().total_seconds() / 60, 1) if soumission.termine_le else '',
            ]
            for question in questions:
                reponse = soumission.reponses.filter(question=question).first()
                if reponse:
                    ligne.append(reponse.obtenir_valeur_affichage())
                else:
                    ligne.append('')
            writer.writerow(ligne)

        return response

    def exporter_excel(self):
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Résultats'

        questions = Question.objects.filter(
            section__sondage=self
        ).order_by('section__ordre', 'ordre')

        entetes = ['Date soumission', 'Durée (min)'] + [q.texte for q in questions]
        ws.append(entetes)

        # Style en-têtes
        for cell in ws[1]:
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(fill_type='solid', fgColor='2563EB')
            cell.alignment = Alignment(horizontal='center')

        for soumission in self.soumission_set.filter(est_complete=True).order_by('termine_le'):
            ligne = [
                soumission.termine_le.strftime('%d/%m/%Y %H:%M') if soumission.termine_le else '',
                round(soumission.obtenir_duree().total_seconds() / 60, 1) if soumission.termine_le else '',
            ]
            for question in questions:
                reponse = soumission.reponses.filter(question=question).first()
                ligne.append(reponse.obtenir_valeur_affichage() if reponse else '')
            ws.append(ligne)

        for col in ws.columns:
            max_len = max((len(str(cell.value or '')) for cell in col), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="sondage_{self.slug}.xlsx"'
        return response


class SectionSondage(models.Model):
    sondage = models.ForeignKey(Sondage, on_delete=models.CASCADE, related_name='sections')
    titre = models.CharField(max_length=200)
    description = models.TextField(blank=True)
    ordre = models.IntegerField(default=0)

    class Meta:
        verbose_name = 'Section'
        verbose_name_plural = 'Sections'
        ordering = ['ordre']

    def __str__(self):
        return f'{self.sondage.titre} — {self.titre}'

    def obtenir_questions(self):
        return self.questions.order_by('ordre')


class Question(models.Model):
    TYPE_CHOIX_UNIQUE = 'choix_unique'
    TYPE_CHOIX_MULTIPLE = 'choix_multiple'
    TYPE_ECHELLE = 'echelle'
    TYPE_TEXTE = 'texte'
    TYPES = [
        (TYPE_CHOIX_UNIQUE, 'Choix unique'),
        (TYPE_CHOIX_MULTIPLE, 'Choix multiple'),
        (TYPE_ECHELLE, 'Échelle'),
        (TYPE_TEXTE, 'Texte libre'),
    ]

    section = models.ForeignKey(SectionSondage, on_delete=models.CASCADE, related_name='questions')
    texte = models.TextField()
    type_question = models.CharField(max_length=20, choices=TYPES)
    est_obligatoire = models.BooleanField(default=True)
    ordre = models.IntegerField(default=0)
    echelle_min = models.IntegerField(default=1)
    echelle_max = models.IntegerField(default=10)
    libelle_min = models.CharField(max_length=100, blank=True)
    libelle_max = models.CharField(max_length=100, blank=True)
    # Logique conditionnelle : afficher cette question seulement si ce choix a été sélectionné
    condition_choix = models.ForeignKey(
        'Choix', null=True, blank=True,
        on_delete=models.SET_NULL, related_name='questions_conditionnelles'
    )

    class Meta:
        verbose_name = 'Question'
        verbose_name_plural = 'Questions'
        ordering = ['ordre']

    def __str__(self):
        return self.texte[:80]

    def obtenir_choix(self):
        return self.choix.order_by('ordre')

    def est_type_choix(self):
        return self.type_question in (self.TYPE_CHOIX_UNIQUE, self.TYPE_CHOIX_MULTIPLE)

    def est_type_echelle(self):
        return self.type_question == self.TYPE_ECHELLE

    def est_type_texte(self):
        return self.type_question == self.TYPE_TEXTE

    def obtenir_nombre_reponses(self):
        from apps.responses.models import Reponse
        return Reponse.objects.filter(
            question=self, soumission__est_complete=True
        ).count()

    def obtenir_taux_reponses(self):
        total = self.section.sondage.soumission_set.filter(est_complete=True).count()
        if total == 0:
            return 0.0
        return round(self.obtenir_nombre_reponses() / total * 100, 1)

    def obtenir_moyenne_echelle(self):
        from apps.responses.models import Reponse
        reponses = Reponse.objects.filter(
            question=self, soumission__est_complete=True, valeur_echelle__isnull=False
        )
        if not reponses.exists():
            return None
        total = sum(r.valeur_echelle for r in reponses)
        return round(total / reponses.count(), 2)

    def obtenir_distribution_reponses(self):
        from apps.responses.models import ReponseChoix
        distribution = {}
        for choix in self.obtenir_choix():
            nb = ReponseChoix.objects.filter(
                choix=choix, reponse__soumission__est_complete=True
            ).count()
            distribution[choix.texte] = nb
        return distribution

    def obtenir_statistiques(self):
        stats = {
            'nb_reponses': self.obtenir_nombre_reponses(),
            'taux_reponses': self.obtenir_taux_reponses(),
        }
        if self.est_type_echelle():
            stats['moyenne'] = self.obtenir_moyenne_echelle()
        if self.est_type_choix():
            stats['distribution'] = self.obtenir_distribution_reponses()
        return stats


class Choix(models.Model):
    question = models.ForeignKey(Question, on_delete=models.CASCADE, related_name='choix')
    texte = models.CharField(max_length=500)
    ordre = models.IntegerField(default=0)

    class Meta:
        verbose_name = 'Choix'
        verbose_name_plural = 'Choix'
        ordering = ['ordre']

    def __str__(self):
        return self.texte

    def obtenir_nombre_selections(self):
        from apps.responses.models import ReponseChoix
        return ReponseChoix.objects.filter(
            choix=self, reponse__soumission__est_complete=True
        ).count()

    def obtenir_pourcentage(self):
        total = self.question.obtenir_nombre_reponses()
        if total == 0:
            return 0.0
        return round(self.obtenir_nombre_selections() / total * 100, 1)


class LienPartage(models.Model):
    sondage = models.ForeignKey(Sondage, on_delete=models.CASCADE, related_name='liens')
    jeton = models.UUIDField(default=uuid.uuid4, unique=True)
    est_public = models.BooleanField(default=True)
    est_actif = models.BooleanField(default=True)
    cree_le = models.DateTimeField(auto_now_add=True)

    class Meta:
        verbose_name = 'Lien de partage'
        verbose_name_plural = 'Liens de partage'

    def __str__(self):
        return f'Lien {self.jeton} — {self.sondage.titre}'

    def obtenir_url_absolue(self):
        from django.conf import settings
        site_url = getattr(settings, 'SITE_URL', 'http://localhost:8000')
        return f'{site_url}/r/{self.jeton}/'

    def desactiver(self):
        self.est_actif = False
        self.save()
